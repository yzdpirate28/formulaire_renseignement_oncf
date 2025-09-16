import os
import random
import csv
from openpyxl import Workbook
from datetime import datetime, timedelta

# Définir le chemin du dossier public
public_dir = os.path.join(os.path.dirname(__file__), "public")
os.makedirs(public_dir, exist_ok=True)

# Champs pour chaque formulaire
fields_catenaire = [
    "date", "dric", "dt", "up", "section", "voie", "doc", "pk_debut", "pk_fin",
    "heure_debut_prevu", "heure_fin_prevu", "delai_prevu", "heure_debut_acc", "heure_fin_acc",
    "nature_travaux", "engin_exploite", "description_travaux", "charge_consignation", "date_renseignement"
]

fields_sousstation = [
    "date", "dric", "district", "sst_ps", "type_intervention", "equipements",
    "charge_exploitation", "doc", "heure_entree", "heure_sortie", "delai_acces",
    "heure_debut_travaux", "heure_fin_travaux", "nature_intervention", "consistance_travaux"
]

# Valeurs d'exemple pour les champs à choix
DRIC = ["SUD", "NORD", "CENTRE"]
DT = ["CAT101", "LC111", "LC112", "LC121", "LC122", "LC211", "LC212", "LC213", "LC221", "LC222", "LC223", "LC311", "LC312", "LC321"]
UP = ["CAT1011", "CAT1012", "CAT1013", "LC1111", "LC1112", "LC1121", "LC1122", "LC1211", "LC1212", "LC1221", "LC1222"]
NATURE_TRAVAUX = ["VSP", "REVISION PROGRAMME", "INTERVENTION PONCTUELLE", "RELEVE DE GEOMETRIE", "M CORRECTIVE"]
ENGIN = ["VMT 1130", "VMT 1131", "VMT 1132", "Donnelli 570+EPA 54", "SANS ENGIN", "AUTRE"]
DISTRICT = ["SS101", "SS111", "SS112", "SS211", "SS221", "SS311", "SS321"]
TYPE_INTER = ["SST", "Télécommande"]
NATURE_INTER = ["VL1", "VL2", "VG", "TS", "M CORR", "INSPECTION"]
VOIE_OPTIONS = ["1", "2", "1,2"]

# Liste de 20 noms de personnes
NOMS_PERSONNES = [
    "Amine El Idrissi", "Sara Benali", "Youssef El Fassi", "Fatima Zahra Lahlou", "Omar Bouzid",
    "Hajar El Mansouri", "Rachid Amrani", "Nadia Bennis", "Khalid El Alaoui", "Imane Chraibi",
    "Mohamed Berrada", "Salma Kabbaj", "Anas El Yousfi", "Meryem El Amrani", "Hamza El Ghazali",
    "Siham El Idrissi", "Adil El Khatib", "Lina El Fassi", "Karim El Mansouri", "Rania El Alaoui"
]

# Fonctions pour générer des données aléatoires

def random_date(start_year=2022, end_year=2024):
    start = datetime(start_year, 1, 1)
    end = datetime(end_year, 12, 31)
    delta = end - start
    random_days = random.randint(0, delta.days)
    return (start + timedelta(days=random_days)).strftime("%d/%m/%Y")

def random_time():
    h = random.randint(0, 23)
    m = random.randint(0, 59)
    return f"{h:02d}:{m:02d}"

def random_float_str(min_val=0, max_val=100, decimals=3):
    return f"{random.uniform(min_val, max_val):.{decimals}f}"

def random_text(length=15):
    return ''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789', k=length))

def random_choice(lst):
    return random.choice(lst)

def random_pk_pair(min_km=0, max_km=100, min_span=0.1, max_span=10):
    pk_debut = random.uniform(min_km, max_km - max_span)
    etendue = random.uniform(min_span, max_span)
    pk_fin = pk_debut + etendue
    return f"{pk_debut:.3f}", f"{pk_fin:.3f}"

# Générer des données pour Caténaire
catenaire_rows = []
for _ in range(200):
    pk_debut, pk_fin = random_pk_pair()
    row = [
        random_date(),
        random_choice(DRIC),
        random_choice(DT),
        random_choice(UP),
        random_text(8),  # section
        random_choice(VOIE_OPTIONS),  # voie
        "AHT",  # doc
        pk_debut,  # pk_debut
        pk_fin,    # pk_fin
        random_time(),  # heure_debut_prevu
        random_time(),  # heure_fin_prevu
        str(random.randint(1, 60)),  # delai_prevu
        random_time(),  # heure_debut_acc
        random_time(),  # heure_fin_acc
        random_choice(NATURE_TRAVAUX),
        random_choice(ENGIN),
        random_text(30),  # description_travaux
        random_choice(NOMS_PERSONNES),  # charge_consignation
        random_date()
    ]
    catenaire_rows.append(row)

# Filtrer les interventions du mois 7 (juillet) pour Caténaire
catenaire_rows_juillet = [row for row in catenaire_rows if row[0][3:5] == "07"]

# Générer des données pour Caténaire Juillet (200 lignes, dates en juillet 2025)
catenaire_rows_juillet = []
for _ in range(200):
    pk_debut, pk_fin = random_pk_pair()
    # Date en juillet 2025
    day = random.randint(1, 31)
    date_juillet = f"{day:02d}/07/2025"
    row = [
        date_juillet,
        random_choice(DRIC),
        random_choice(DT),
        random_choice(UP),
        random_text(8),  # section
        random_choice(VOIE_OPTIONS),  # voie
        "AHT",  # doc
        pk_debut,  # pk_debut
        pk_fin,    # pk_fin
        random_time(),  # heure_debut_prevu
        random_time(),  # heure_fin_prevu
        str(random.randint(1, 60)),  # delai_prevu
        random_time(),  # heure_debut_acc
        random_time(),  # heure_fin_acc
        random_choice(NATURE_TRAVAUX),
        random_choice(ENGIN),
        random_text(30),  # description_travaux
        random_choice(NOMS_PERSONNES),  # charge_consignation
        random_date()
    ]
    catenaire_rows_juillet.append(row)

# Générer des données pour Sous-Station
sousstation_rows = []
for _ in range(200):
    row = [
        random_date(),
        random_choice(DRIC),
        random_choice(DISTRICT),
        random_text(6),  # sst_ps
        random_choice(TYPE_INTER),
        random_text(10),  # equipements
        random_choice(NOMS_PERSONNES),  # charge_exploitation
        "AHT",  # doc
        random_time(),  # heure_entree
        random_time(),  # heure_sortie
        str(random.randint(1, 60)),  # delai_acces
        random_time(),  # heure_debut_travaux
        random_time(),  # heure_fin_travaux
        random_choice(NATURE_INTER),
        random_text(30)  # consistance_travaux
    ]
    sousstation_rows.append(row)

# Filtrer les interventions du mois 7 (juillet) pour Sous-Station
sousstation_rows_juillet = [row for row in sousstation_rows if row[0][3:5] == "07"]

# Générer des données pour Sous-Station Juillet (200 lignes, dates en juillet 2025)
sousstation_rows_juillet = []
for _ in range(200):
    # Date en juillet 2025
    day = random.randint(1, 31)
    date_juillet = f"{day:02d}/07/2025"
    row = [
        date_juillet,
        random_choice(DRIC),
        random_choice(DISTRICT),
        random_text(6),  # sst_ps
        random_choice(TYPE_INTER),
        random_text(10),  # equipements
        random_choice(NOMS_PERSONNES),  # charge_exploitation
        "AHT",  # doc
        random_time(),  # heure_entree
        random_time(),  # heure_sortie
        str(random.randint(1, 60)),  # delai_acces
        random_time(),  # heure_debut_travaux
        random_time(),  # heure_fin_travaux
        random_choice(NATURE_INTER),
        random_text(30)  # consistance_travaux
    ]
    sousstation_rows_juillet.append(row)

# Création du fichier Excel pour Caténaire
wb_cat = Workbook()
ws_cat = wb_cat.active
ws_cat.title = "Catenaire"
ws_cat.append(fields_catenaire)
for row in catenaire_rows:
    ws_cat.append(row)
cat_file = os.path.join(public_dir, "formulaire_catenaire2.xlsx")
wb_cat.save(cat_file)

# Création du fichier Excel pour Caténaire Juillet
wb_cat_juillet = Workbook()
ws_cat_juillet = wb_cat_juillet.active
ws_cat_juillet.title = "Catenaire_Juillet"
ws_cat_juillet.append(fields_catenaire)
for row in catenaire_rows_juillet:
    ws_cat_juillet.append(row)
cat_file_juillet = os.path.join(public_dir, "formulaire_catenaire_juillet.xlsx")
wb_cat_juillet.save(cat_file_juillet)

# Création du fichier Excel pour Sous-Station
wb_sous = Workbook()
ws_sous = wb_sous.active
ws_sous.title = "SousStation"
ws_sous.append(fields_sousstation)
for row in sousstation_rows:
    ws_sous.append(row)
sous_file = os.path.join(public_dir, "formulaire_sousstation2.xlsx")
wb_sous.save(sous_file)

# Création du fichier Excel pour Sous-Station Juillet
wb_sous_juillet = Workbook()
ws_sous_juillet = wb_sous_juillet.active
ws_sous_juillet.title = "SousStation_Juillet"
ws_sous_juillet.append(fields_sousstation)
for row in sousstation_rows_juillet:
    ws_sous_juillet.append(row)
sous_file_juillet = os.path.join(public_dir, "formulaire_sousstation_juillet.xlsx")
wb_sous_juillet.save(sous_file_juillet)

# Création du CSV pour Caténaire
cat_csv = os.path.join(public_dir, "formulaire_catenaire2.csv")
with open(cat_csv, mode="w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(fields_catenaire)
    writer.writerows(catenaire_rows)

# Création du CSV pour Caténaire Juillet
cat_csv_juillet = os.path.join(public_dir, "formulaire_catenaire_juillet.csv")
with open(cat_csv_juillet, mode="w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(fields_catenaire)
    writer.writerows(catenaire_rows_juillet)

# Création du CSV pour Sous-Station
sous_csv = os.path.join(public_dir, "formulaire_sousstation2.csv")
with open(sous_csv, mode="w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(fields_sousstation)
    writer.writerows(sousstation_rows)

# Création du CSV pour Sous-Station Juillet
sous_csv_juillet = os.path.join(public_dir, "formulaire_sousstation_juillet.csv")
with open(sous_csv_juillet, mode="w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(fields_sousstation)
    writer.writerows(sousstation_rows_juillet)

print("Fichiers Excel et CSV créés dans le dossier public !")